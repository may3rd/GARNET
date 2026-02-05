import os
import glob
import pandas as pd
import cv2
import garnet.Settings as Settings
from sahi import AutoDetectionModel
from sahi.predict import get_sliced_prediction
import easyocr
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

"""
This script performs object detection on images using a specified model,
applies OCR to detected symbols that require text extraction,
filters detections by specified class keywords,
saves cropped detected objects,
and exports the results to an Excel file with one sheet per image.
The Excel file includes embedded cropped images for each detection.
"""

# Initialize OCR reader with English language
reader = easyocr.Reader(["en"])

# Characters allowed in OCR recognition (capital letters, digits, dash, quote)
allowlist = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-"'

# Kernel used for morphological operations in OCR preprocessing
kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))

# Load symbol settings, including which symbols require OCR extraction
settings = Settings.Settings()
SYMBOL_WITH_TEXT = set(settings.SYMBOL_WITH_TEXT)


def run_prediction_and_export_excel(
    image_path,
    model_type="yolov8",
    model_path="../yolo_weights/yolo11n_PPCL_640_20250204.pt",
    output_path="predictions",
    model_confidence_threshold=0.5,
    image_size=640,
    overlab_ratio=0.5,
    preform_standard_pred=True,
    postprocess_type="GREEDYNMM",
    postprocess_match_metric="IOU",
    postprocess_match_threshold=0.2,
    postprocess_class_agnostic=True,
    class_list=None,
    rect_th=2,
    text_size=0.3,
    hide_conf=True,
    hide_labels=True,
):
    """
    Run object detection prediction on images (single or folder), perform OCR on specific detected symbols,
    filter detections by class list if provided, save cropped detected objects,
    export prediction results including embedded crops into an Excel file with one sheet per image.

    Args:
        image_path (str): Path to a single image file or directory containing images.
        model_type (str): Type of detection model to use (default: 'yolov8').
        model_path (str): Path to the model weights file.
        output_path (str): Directory to save prediction outputs and Excel file.
        model_confidence_threshold (float): Minimum confidence threshold for detections.
        image_size (int): Image slice size for prediction model.
        overlab_ratio (float): Overlap ratio for slicing images during prediction.
        preform_standard_pred (bool): Whether to perform standard full image prediction.
        postprocess_type (str): Postprocessing algorithm type.
        postprocess_match_metric (str): Metric for matching boxes in postprocessing.
        postprocess_match_threshold (float): Threshold for matching boxes.
        postprocess_class_agnostic (bool): Whether postprocessing is class agnostic.
        class_list (list[str], optional): List of class filter keywords for detection filtering.
        rect_th (int): Rectangle thickness for drawing bounding boxes.
        text_size (float): Text size for visualization labels.
        hide_conf (bool): Whether to hide confidence scores on visualizations.
        hide_labels (bool): Whether to hide labels on visualizations.

    Returns:
        None
    """

    if class_list is None:
        class_list = []

    # Create output directory if it does not exist
    os.makedirs(output_path, exist_ok=True)

    # Prepare list of image paths to process: support single image or directory (recursive)
    image_paths = []
    if os.path.isfile(image_path):
        # Single image input
        image_paths = [image_path]
    elif os.path.isdir(image_path):
        # Search recursively for image files in directory
        for ext in ("*.png", "*.jpg", "*.jpeg", "*.bmp", "*.tif", "*.tiff"):
            image_paths.extend(glob.glob(os.path.join(image_path, "**", ext), recursive=True))
    else:
        print(f"ERROR: Path '{image_path}' is not a file or directory.")
        return

    if not image_paths:
        print("No images found.")
        return

    image_paths.sort()
    total_images = len(image_paths)
    excel_data = {}

    # Directory to save cropped detected objects
    crop_save_dir = os.path.join(output_path, "crops")
    os.makedirs(crop_save_dir, exist_ok=True)

    # Prepare model category mapping if using ONNX yolov8 model
    category_mapping = None
    if model_type == "yolov8onnx":
        import onnx
        import ast
        model = onnx.load(model_path)
        props = {p.key: p.value for p in model.metadata_props}
        names = ast.literal_eval(props["names"])
        category_mapping = {str(key): value for key, value in names.items()}

    # Load detection model with given parameters
    detection_model = AutoDetectionModel.from_pretrained(
        model_type=model_type,
        model_path=model_path,
        confidence_threshold=model_confidence_threshold,
        category_mapping=category_mapping,
    )

    # Convert class filter terms to lowercase for case-insensitive matching
    search_terms = [term.lower() for term in class_list]

    def match_label(label, search_terms):
        """
        Check if any search term is a substring of the label (case-insensitive).
        """
        label = label.lower()
        return any(term in label for term in search_terms)

    for idx, img_path in enumerate(image_paths):
        print(f"Processing {idx + 1}/{total_images}: {img_path}")

        # Load image using OpenCV (BGR format)
        img_cv = cv2.imread(img_path)
        if img_cv is None:
            print(f"Warning: Failed to read image {img_path}, skipping.")
            continue

        # Convert BGR image to RGB for model prediction
        img_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
        h, w = img_rgb.shape[:2]

        # Run sliced prediction on image with the detection model
        result = get_sliced_prediction(
            image=img_rgb,
            detection_model=detection_model,
            perform_standard_pred=preform_standard_pred,
            slice_height=image_size,
            slice_width=image_size,
            overlap_height_ratio=overlab_ratio,
            postprocess_type=postprocess_type,
            postprocess_match_metric=postprocess_match_metric,
            postprocess_match_threshold=postprocess_match_threshold,
            postprocess_class_agnostic=postprocess_class_agnostic,
            verbose=2,
        )

        # Export prediction visualizations (bounding boxes and labels) to output directory
        image_name = os.path.splitext(os.path.basename(img_path))[0]
        result.export_visuals(
            export_dir=output_path,
            file_name=image_name,
            text_size=text_size,
            rect_th=rect_th,
            hide_conf=hide_conf,
            hide_labels=hide_labels
        )

        # Filter predictions if class_list specified, otherwise keep all
        if class_list:
            filtered_predictions = [
                pred for pred in result.object_prediction_list
                if match_label(pred.category.name, search_terms)
            ]
        else:
            filtered_predictions = result.object_prediction_list

        # Prepare rows for Excel export, including OCR text and cropped image saving
        excel_rows = []

        for pred in filtered_predictions:
            bbox = pred.bbox.to_xywh()
            x, y, width, height = map(int, bbox)

            # Save cropped image of detected object for embedding into Excel
            crop_file = ""
            if width > 0 and height > 0:
                crop = img_cv[y:y + height, x:x + width]
                crop_file = os.path.join(
                    crop_save_dir,
                    f"{image_name}_{pred.category.name}_{x}_{y}.png"
                )
                cv2.imwrite(crop_file, crop)

            # Perform OCR if symbol category requires text extraction
            if pred.category.name in SYMBOL_WITH_TEXT:
                cropped = img_cv[y:y + height, x:x + width]
                # Preprocess crop image for OCR by inverting colors and dilating
                processed = cv2.bitwise_not(
                    cv2.dilate(
                        cv2.bitwise_not(cropped),
                        kernel,
                        iterations=1
                    )
                )
                ocr_result = reader.readtext(
                    processed,
                    decoder="wordbeamsearch",
                    batch_size=8,
                    paragraph=True,
                    detail=0,
                    text_threshold=0.7,
                    allowlist=allowlist,
                )
                text = " ".join(str(item) for item in ocr_result) if ocr_result else ""
            else:
                text = ""

            excel_rows.append({
                "category id": pred.category.id,
                "label": pred.category.name,
                "confidence": pred.score.value,
                "x": x,
                "y": y,
                "width": width,
                "height": height,
                "extracted_text": text,
                "ocr_confidence": 1.0 if text else 0.0,
                "cropped_img": crop_file,  # Path to cropped image for Excel embedding
            })

        # Sort detections by category id and confidence for consistent Excel ordering
        excel_rows.sort(key=lambda x: (x["category id"], x["confidence"]))

        # Excel sheet names have max length 31; sanitize sheet name from image file name
        sheet_name = os.path.basename(img_path)[:31].replace(":", "_").replace("/", "_")

        # Store detection data per image sheet
        excel_data[sheet_name] = excel_rows

    # Export all detections to Excel with embedded cropped images
    save_predictions_to_excel(excel_data, os.path.join(output_path, "predictions.xlsx"))


def save_predictions_to_excel(excel_data: dict, output_path: str):
    """
    Save detection results to an Excel workbook with one sheet per image.
    Then embed the cropped detection images into the Excel cells at actual size.

    Args:
        excel_data (dict): Dictionary with sheet_name as key and list of detection dicts as values.
        output_path (str): File path to save the Excel workbook.

    Returns:
        None
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Write detection data to Excel sheets (without images)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows in excel_data.items():
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Excel file saved to {output_path}")

    # Load workbook again to embed cropped images into cells
    wb = load_workbook(output_path)
    for sheet_name, rows in excel_data.items():
        ws = wb[sheet_name]
        if not rows:
            continue

        # Identify column for cropped image embedding
        header = list(rows[0].keys())
        try:
            crop_col_idx = header.index("cropped_img") + 1  # Excel columns are 1-based
        except ValueError:
            crop_col_idx = len(header)  # Default to last column if not found

        crop_col_letter = get_column_letter(crop_col_idx)

        # Embed each cropped image into corresponding cell, adjust row height and column width
        for i, row in enumerate(rows, start=2):  # Data starts from row 2 (row 1 = headers)
            img_path = row.get("cropped_img", "")
            if img_path and os.path.isfile(img_path):
                try:
                    with PILImage.open(img_path) as pil_img:
                        orig_w, orig_h = pil_img.size

                    img = XLImage(img_path)
                    img.width, img.height = orig_w, orig_h

                    # Adjust column width and row height to fit image size
                    ws.column_dimensions[crop_col_letter].width = max(
                        ws.column_dimensions[crop_col_letter].width or 0, orig_w / 7
                    )
                    ws.row_dimensions[i].height = max(
                        ws.row_dimensions[i].height or 0, orig_h * 0.75
                    )

                    cell = f"{crop_col_letter}{i}"
                    ws.add_image(img, cell)
                except Exception as e:
                    print(f"Warning: could not embed image {img_path}: {e}")

    wb.save(output_path)
    print(f"Embedded all cropped images into {output_path}")


def parse_args():
    """
    Parse command line arguments for running the prediction and export script.

    Returns:
        argparse.Namespace: Parsed arguments namespace.
    """
    import argparse
    parser = argparse.ArgumentParser(description="Run prediction and export to Excel")
    parser.add_argument("--image_path", type=str, default="test/ppcl", help="Path to image file or directory")
    parser.add_argument("--model_type", type=str, default="yolov8onnx", help="Model type for detection")
    parser.add_argument("--model_path", type=str, default="yolo_weights/yolo11s_PPCL_640_20250207.onnx", help="Path to model weights")
    parser.add_argument("--output_path", type=str, default="predictions/ppcl", help="Directory for outputs")
    parser.add_argument("--conf", type=float, default=0.8, help="Confidence threshold for detections")
    parser.add_argument("--image_size", type=int, default=640, help="Image slice size for prediction")
    parser.add_argument("--overlab_ratio", type=float, default=0.2, help="Overlap ratio for sliced prediction")
    parser.add_argument("--preform_standard_pred", type=bool, default=True, help="Perform standard full image prediction")
    parser.add_argument("--postprocess_type", type=str, default="GREEDYNMM", help="Postprocessing method type")
    parser.add_argument("--postprocess_match_metric", type=str, default="IOS", help="Postprocessing match metric")
    parser.add_argument("--postprocess_match_threshold", type=float, default=0.1, help="Postprocessing match threshold")
    parser.add_argument("--postprocess_class_agnostic", type=bool, default=True, help="Postprocessing class agnostic flag")
    parser.add_argument("--class_list", nargs="+", default=["valve"], help="List of class filter keywords")
    parser.add_argument("--rect_th", type=int, default=4, help="Rectangle thickness for visualization")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()

    # Normalize class_list if input as single comma-separated string
    if len(args.class_list) == 1 and "," in args.class_list[0]:
        args.class_list = [s.strip() for s in args.class_list[0].split(",")]

    run_prediction_and_export_excel(
        image_path=args.image_path,
        model_type=args.model_type,
        model_path=args.model_path,
        output_path=args.output_path,
        model_confidence_threshold=args.conf,
        image_size=args.image_size,
        overlab_ratio=args.overlab_ratio,
        preform_standard_pred=args.preform_standard_pred,
        postprocess_type=args.postprocess_type,
        postprocess_match_metric=args.postprocess_match_metric,
        postprocess_match_threshold=args.postprocess_match_threshold,
        postprocess_class_agnostic=args.postprocess_class_agnostic,
        class_list=args.class_list,
        rect_th=args.rect_th,
        hide_conf=True,
        hide_labels=True
    )
